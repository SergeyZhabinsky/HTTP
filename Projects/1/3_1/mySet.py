# import xlrd, xlwt
import openpyxl
# rb = xlrd.open_workbook('D:\\temp\\1.xlsx', formatting_info=True)
# D:\temp\1.xlsx
# workbook = xlrd.open_workbook("D:\\temp\\1.xlsx")
wookbook = openpyxl.load_workbook("D:\\temp\\1.xlsx")
worksheet = wookbook.active

for i in range(0, worksheet.max_row):
    for col in worksheet.iter_cols(1, worksheet.max_column):
        print(col[i].value, end="\t\t")
    print('')